import os
import sys
from src import utils
import pandas as pd
from src.logger import logging
from src.exception import ExcelException
from src.entity import entity_artifact, entity_config
from src.components import data_ingestion, data_validation
from src import properties
import openpyxl as xl
from openpyxl.utils.dataframe import dataframe_to_rows
import win32com.client


def refresh_excel_file(excel_file_path):
    try:
        logging.info('Inside function: refresh_excel_file')
        logging.info(f'File: {excel_file_path} refresh initiated')
        file = win32com.client.Dispatch('Excel.Application')
        file.Visible = 0
        book = file.Workbooks.open(excel_file_path)
        book.RefreshAll()
        book.Save()
        file.Quit()
        logging.info(f'File: {excel_file_path} refresh completed')

    except Exception as e:
        logging.error(e)
        raise ExcelException(e, sys)


class DataCalculation:
    def __init__(self, data_validation_artifact: entity_artifact.DataValidationArtifact,
                 data_calculation_config: entity_config.DataCalculationConfig):
        try:
            logging.info(f"{'>>' * 20} Inside class: DataCalculation {'<<' * 20}")
            self.data_validation_artifact = data_validation_artifact
            self.data_calculation_config = data_calculation_config
            logging.info('DataValidation Initiation completed')
        except Exception as e:
            raise ExcelException(e, sys)

    def execute(self):
        try:
            logging.info(f"{'>>' * 20} Inside Method: DataCalculation.execute {'<<' * 20}")

            os.makedirs(self.data_calculation_config.data_calculation_dir, exist_ok=True)
            logging.info(f'Created directory: {self.data_calculation_config.data_calculation_dir}')

            input_path_list = self.data_validation_artifact.valid_input_path_list
            logging.info(f'Input file path list: {input_path_list}')

            template_path = self.data_validation_artifact.valid_template_path
            logging.info(f'Template path list: {template_path}')

            input_file_list = [os.path.basename(path) for path in input_path_list]
            logging.info(f'Input file name list: {input_file_list}')

            sheet_mapping = properties.SheetMapping.sheet_mapping
            # load template workbook
            template_workbook = xl.load_workbook(filename=template_path)

            for file in input_path_list:
                mapping = sheet_mapping[os.path.basename(file)]
                input_workbook = xl.load_workbook(filename=file, keep_links=False)
                input_sheet_list = input_workbook.sheetnames
                for sheet in input_sheet_list:
                    input_df = pd.read_excel(io=file, sheet_name=sheet)
                    input_rows = dataframe_to_rows(input_df, index=False, header=True)
                    template_sheet = template_workbook[mapping[sheet]]
                    logging.info(f'In Template file, sheet {template_sheet} process started')
                    for r_idx, row in enumerate(input_rows, 1):
                        for c_idx, value in enumerate(row, 1):
                            try:
                                template_sheet.cell(row=r_idx, column=c_idx, value=float(value))
                            except:
                                template_sheet.cell(row=r_idx, column=c_idx, value=value)
                    logging.info(f'In Template file, sheet {template_sheet} process completed')
            template_workbook.save(template_path)
            logging.info('Data saved in template file')

            # Refreshing Excel file
            refresh_excel_file(excel_file_path=template_path)

            # Temp output file path generation
            temp_output_path = os.path.join(self.data_calculation_config.data_calculation_dir, 'temp_output.xlsx')
            logging.info(f'Temporary output path created: {temp_output_path}')

            # Temp output file creation
            utils.file_operation(source_file=template_path, destination_file=temp_output_path, operation='move')

            data_calculation_artifact = entity_artifact.DataCalculationArtifact(temp_output_file=temp_output_path)
            logging.info(f"Data Calculation artifact created: {data_calculation_artifact}")
            logging.info(f"{'>>' * 20} Exit Method: DataCalculation.execute {'<<' * 20}")

            return data_calculation_artifact

        except Exception as e:
            raise ExcelException(e, sys)
